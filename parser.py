"""
File parser for CSV and Excel uploads.
Columns A-G: Materialnummer, Preis, Preis_ab, MWST_in_, Preis_2, EUR (Währung), Territory
"""
import csv
import io
from datetime import datetime, date
from typing import Optional

from openpyxl import load_workbook

# If the first cell of row 1 matches any of these (case-insensitive),
# the row is treated as a header and skipped.
HEADER_KEYWORDS = {"materialnummer", "material", "matnr", "artikelnummer"}


def parse_file(content: bytes, filename: str) -> list[dict]:
    """Parse uploaded file (CSV or Excel) and return a list of row dicts."""
    name_lower = (filename or "").lower()
    if name_lower.endswith(".csv"):
        return _parse_csv(content)
    elif name_lower.endswith((".xlsx", ".xls")):
        return _parse_excel(content)
    else:
        raise ValueError(
            f"Nicht unterstütztes Dateiformat: '{filename}'. "
            "Erlaubt sind .xlsx, .xls und .csv."
        )


# ---------------------------------------------------------------------------
# CSV
# ---------------------------------------------------------------------------

def _parse_csv(content: bytes) -> list[dict]:
    # Try UTF-8 with BOM first, fall back to Latin-1
    for encoding in ("utf-8-sig", "utf-8", "latin-1"):
        try:
            text = content.decode(encoding)
            break
        except UnicodeDecodeError:
            continue
    else:
        raise ValueError("CSV-Datei konnte nicht dekodiert werden (UTF-8 oder Latin-1 erwartet).")

    reader = csv.reader(io.StringIO(text), delimiter=";")
    rows = list(reader)

    if not rows:
        return []

    start_idx = _detect_header(rows[0][0] if rows[0] else "")
    result = []
    for row in rows[start_idx:]:
        if not any(cell.strip() for cell in row):
            continue  # skip empty rows
        result.append(_build_row_dict([cell.strip() for cell in row]))
    return result


# ---------------------------------------------------------------------------
# Excel
# ---------------------------------------------------------------------------

def _parse_excel(content: bytes) -> list[dict]:
    wb = load_workbook(io.BytesIO(content), data_only=True, read_only=True)
    ws = wb.active

    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    if not rows:
        return []

    first_cell = str(rows[0][0] or "").strip()
    start_idx = _detect_header(first_cell)

    result = []
    for row in rows[start_idx:]:
        str_row = [_cell_to_string(v, col_idx) for col_idx, v in enumerate(row)]
        if not any(v.strip() for v in str_row):
            continue  # skip empty rows
        result.append(_build_row_dict(str_row))
    return result


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _detect_header(first_cell: str) -> int:
    """Return 1 if first row looks like a header, else 0."""
    return 1 if first_cell.strip().lower() in HEADER_KEYWORDS else 0


def _cell_to_string(value, col_idx: int) -> str:
    """Convert an openpyxl cell value to a normalised string."""
    if value is None:
        return ""

    # Date / datetime → dd.MM.yyyy
    if isinstance(value, (datetime, date)):
        if isinstance(value, datetime):
            return value.strftime("%d.%m.%Y")
        return value.strftime("%d.%m.%Y")

    # Numeric columns: Preis (1), Preis_2 (4)
    if isinstance(value, float):
        # Avoid floating-point representation issues (e.g. 19.130000001)
        # by using Python's str() which gives the shortest round-trip repr.
        s = str(value)
        # If it's a whole number, drop the decimal part
        if value == int(value):
            return str(int(value))
        return s

    if isinstance(value, int):
        return str(value)

    return str(value).strip()


def _build_row_dict(cells: list[str]) -> dict:
    """Map a list of cell strings to the canonical row dict."""

    def get(idx: int) -> str:
        return cells[idx].strip() if idx < len(cells) else ""

    return {
        "Materialnummer": get(0),
        "Preis": get(1),
        "Preis_ab": get(2),
        "MWST_in_": get(3),
        "Preis_2": get(4),
        "EUR": get(5),       # Währung → JSON key "EUR"
        "Territory": get(6),
    }
